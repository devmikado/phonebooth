from django.http import HttpResponse
from openpyxl import *
import openpyxl
import io
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK, BORDER_MEDIUM
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series


# Worksheet style
def style(ws, fieldnames, report_type):
    font = Font(name='Calibri',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    font2 = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    color = Font(name='Calibri',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FFFFFF')
    border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000')
        )

    thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )
    fill=PatternFill(start_color = '00C0C0C0',
            end_color = '00C0C0C0',
            fill_type = 'solid')

    fill2=PatternFill(start_color = 'ebd3ae',
            end_color = 'ebd3ae',
            fill_type = 'solid')
    
    fill3=PatternFill(start_color = 'f6fcc0',
            end_color = 'f6fcc0',
            fill_type = 'solid')

    fill4=PatternFill(start_color = 'ffffff',
            end_color = 'ffffff',
            fill_type = 'solid')

    dd = Font(underline='single', color='000000FF')
    # row_count = ws.max_row
    # column_count = ws.max_column

    for cell1 in ws['1:1']:
        # cell2.fill = fill
        cell1.font = font2
    
    for cell2 in ws['2:2']:
        cell2.fill = fill
        cell2.font = font  

    ws.row_dimensions[2].height = 40
    # ws.row_dimensions[1].height = 40

    for row in ws:
        for cell1 in row:
            cell1.border = border
            cell1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    
    column_widths = []
    for col in fieldnames:
        for i in range(len(col)):
            if len(column_widths) > i:
                if len(col) > column_widths[i]:
                    column_widths[i] = len(col)
            else:
                column_widths += [len(col)]
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = 20
    
    # ws.column_dimensions[get_column_letter(3)].width = 30
    # ws.column_dimensions[get_column_letter(5)].width = 40

    

def write_excel(request, data, report_type):
    wb = Workbook()
    ws1 = wb.active
    ws2 = wb.active
    
    if report_type == 3:
        ws1.title = "Report-3"
        ws1.merge_cells('A1:B1')
        n1 = ws1.cell(row=1,column=1)
        n1.value = "Location wise happiness quotient"
        
        fieldnames = ["Ref.No.", "Date", "Weekday", "Location", "Culture", "Happy Comments Percentage", "Negative Comments Percentage", "Neutral Comments Percentage"]
        ws1.append(fieldnames)
        count = 0
        
        for i in data['dataRow']:
            count += 1
            issue_data = [[
                    count,
                    i['date'],
                    i['weekday'],
                    i['location'],
                    i['culture'],
                    i['happy_comments'],
                    i['negative_comments'],
                    i['neutral_comments']
                ]]
            for data in issue_data:
                ws1.append(data)
    
        style(ws1,fieldnames, report_type)

    if report_type == 4:
        ws2.title = "Report-4"
        ws2.merge_cells('A1:B1')
        n1 = ws2.cell(row=1,column=1)
        n1.value = "Location wise languages comment counts"
        
        fieldnames = ["Ref.No.", "Date", "Weekday", "Location", "Culture", "Language", "Counts"]
        ws2.append(fieldnames)
        count = 0
        
        for i in data['dataRow']:
            count += 1
            issue_data = [[
                    count,
                    i['date'],
                    i['weekday'],
                    i['location'],
                    i['culture'],
                    i['language'],
                    i['comments_count'],
                ]]
            for data in issue_data:
                ws2.append(data)

        style(ws2, fieldnames, report_type)

    wb.close()

    
    return wb